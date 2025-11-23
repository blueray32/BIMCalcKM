"""
Export utilities for executive dashboards.

Provides CSV and Excel export functionality for dashboard metrics.
"""

import csv
import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _sanitize_sheet_name(name: str) -> str:
    """Ensure Excel sheet name is valid and within length."""
    safe = "".join("-" if ch in '[]:*?/\\' else ch for ch in name).strip()
    if not safe:
        safe = "Sheet"
    return safe[:31]


def format_currency(value: float | None, currency: str = "EUR") -> str:
    """Format currency value for export."""
    if value is None:
        return "N/A"
    return f"{currency} {value:,.2f}"


def format_percentage(value: float | None) -> str:
    """Format percentage value for export."""
    if value is None:
        return "N/A"
    return f"{value:.1f}%"


def format_count(value: int | None) -> str:
    """Format count value for export."""
    if value is None:
        return "0"
    return str(value)


class ExcelExporter:
    """Excel workbook exporter with styling."""

    def __init__(self, title: str, org_id: str, project_id: str):
        self.wb = Workbook()
        self.title = title
        self.org_id = org_id
        self.project_id = project_id
        self.timestamp = datetime.now()

        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        # Define styles
        self.header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def add_metadata_sheet(self):
        """Add metadata sheet with export information."""
        ws = self.wb.create_sheet(_sanitize_sheet_name("Export Info"), 0)

        # Title
        ws['A1'] = self.title
        ws['A1'].font = Font(bold=True, size=16, color="667EEA")

        # Metadata
        ws['A3'] = "Organization:"
        ws['B3'] = self.org_id
        ws['A4'] = "Project:"
        ws['B4'] = self.project_id
        ws['A5'] = "Generated:"
        ws['B5'] = self.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        ws['A6'] = "System:"
        ws['B6'] = "BIMCalc Executive Dashboard"

        # Style metadata
        for row in range(3, 7):
            ws[f'A{row}'].font = Font(bold=True)

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def add_kpi_sheet(self, name: str, data: list[dict[str, Any]]):
        """Add a sheet with KPI data."""
        ws = self.wb.create_sheet(_sanitize_sheet_name(name))

        if not data:
            ws['A1'] = "No data available"
            return

        # Headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # Data rows
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        # Freeze header row
        ws.freeze_panes = 'A2'

    def save(self) -> bytes:
        """Save workbook to bytes."""
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def export_dashboard_to_excel(metrics, org_id: str, project_id: str) -> bytes:
    """Export dashboard metrics to Excel."""
    exporter = ExcelExporter("Executive Dashboard", org_id, project_id)
    exporter.add_metadata_sheet()

    # KPI Summary
    kpi_data = [
        {
            "Metric": "Total Cost (Net)",
            "Value": format_currency(metrics.total_cost_net, metrics.currency)
        },
        {
            "Metric": "Total Cost (Gross)",
            "Value": format_currency(metrics.total_cost_gross, metrics.currency)
        },
        {
            "Metric": "High Risk Cost",
            "Value": format_currency(metrics.high_risk_cost, metrics.currency)
        },
        {
            "Metric": "Total Items",
            "Value": format_count(metrics.total_items)
        },
        {
            "Metric": "Matched Items",
            "Value": format_count(metrics.matched_items)
        },
        {
            "Metric": "Match Coverage",
            "Value": format_percentage(metrics.match_percentage)
        },
        {
            "Metric": "Auto-Approval Rate",
            "Value": format_percentage(metrics.auto_approval_rate)
        },
        {
            "Metric": "Average Confidence",
            "Value": format_percentage(metrics.avg_confidence) if metrics.avg_confidence else "N/A"
        },
        {
            "Metric": "High Confidence %",
            "Value": format_percentage(metrics.high_confidence_percentage)
        },
        {
            "Metric": "Pending Review",
            "Value": format_count(metrics.total_pending_review)
        },
        {
            "Metric": "High Urgency Items",
            "Value": format_count(metrics.high_urgency_count)
        },
        {
            "Metric": "Health Score",
            "Value": f"{metrics.health_score}/100 ({metrics.health_status})"
        }
    ]
    exporter.add_kpi_sheet("KPI Summary", kpi_data)

    # Recent Activity
    activity_data = [
        {
            "Activity": "Recent Matches (7 days)",
            "Count": format_count(metrics.recent_matches)
        },
        {
            "Activity": "Recent Approvals (7 days)",
            "Count": format_count(metrics.recent_approvals)
        },
        {
            "Activity": "Recent Ingestions (7 days)",
            "Count": format_count(metrics.recent_ingestions)
        }
    ]
    exporter.add_kpi_sheet("Recent Activity", activity_data)

    classification_data = getattr(metrics, "classification_distribution", None)
    if classification_data:
        class_rows = [
            {
                "Classification": f"{cls.get('code', 'N/A')} - {cls.get('name', 'N/A')}",
                "Items": format_count(cls.get('items', 0)),
                "Matched Items": format_count(cls.get('matched', 0)),
                "Matched Cost": format_currency(cls.get('matched_cost', 0), metrics.currency),
                "Cost Share": format_percentage(cls.get('cost_share', 0)),
                "Avg Confidence": format_percentage(cls.get('avg_confidence')) if cls.get('avg_confidence') else "N/A",
            }
            for cls in classification_data
        ]
        exporter.add_kpi_sheet("Top Classifications", class_rows)

        for cls in classification_data:
            sheet_name = f"{cls.get('code', 'Class')}"
            detail_rows = [
                {"Metric": "Classification", "Value": f"{cls.get('code', 'N/A')} - {cls.get('name', 'N/A')}"},
                {"Metric": "Total Items", "Value": format_count(cls.get('items', 0))},
                {"Metric": "Matched Items", "Value": format_count(cls.get('matched', 0))},
                {"Metric": "Matched Cost", "Value": format_currency(cls.get('matched_cost', 0), metrics.currency)},
                {"Metric": "Cost Share", "Value": format_percentage(cls.get('cost_share', 0))},
                {"Metric": "Average Confidence", "Value": format_percentage(cls.get('avg_confidence')) if cls.get('avg_confidence') else "N/A"},
            ]
            exporter.add_kpi_sheet(f"{sheet_name} Detail", detail_rows)

    return exporter.save()


def export_progress_to_excel(metrics, org_id: str, project_id: str) -> bytes:
    """Export progress metrics to Excel."""
    exporter = ExcelExporter("Match Progress Report", org_id, project_id)
    exporter.add_metadata_sheet()

    # Summary
    unmatched_items = metrics.total_items - metrics.matched_items
    match_percentage = (metrics.matched_items / metrics.total_items * 100) if metrics.total_items > 0 else 0

    summary_data = [
        {
            "Metric": "Total Items",
            "Value": format_count(metrics.total_items)
        },
        {
            "Metric": "Matched Items",
            "Value": format_count(metrics.matched_items)
        },
        {
            "Metric": "Unmatched Items",
            "Value": format_count(unmatched_items)
        },
        {
            "Metric": "Match Coverage",
            "Value": format_percentage(match_percentage)
        },
        {
            "Metric": "Pending Review",
            "Value": format_count(metrics.pending_review)
        },
        {
            "Metric": "Auto-Approved",
            "Value": format_count(metrics.auto_approved)
        },
        {
            "Metric": "Flagged Critical",
            "Value": format_count(metrics.flagged_critical)
        },
        {
            "Metric": "Flagged Advisory",
            "Value": format_count(metrics.flagged_advisory)
        },
        {
            "Metric": "Overall Completion",
            "Value": format_percentage(metrics.overall_completion)
        },
        {
            "Metric": "Overall Status",
            "Value": metrics.overall_status
        }
    ]
    exporter.add_kpi_sheet("Match Summary", summary_data)

    # Confidence Distribution
    confidence_data = [
        {
            "Confidence Level": "High (≥85%)",
            "Count": format_count(metrics.confidence_high)
        },
        {
            "Confidence Level": "Medium (70-84%)",
            "Count": format_count(metrics.confidence_medium)
        },
        {
            "Confidence Level": "Low (<70%)",
            "Count": format_count(metrics.confidence_low)
        }
    ]
    exporter.add_kpi_sheet("Confidence Distribution", confidence_data)

    # Classification Coverage
    if metrics.classification_coverage:
        class_data = [
            {
                "Classification Code": cls.get('code', 'N/A'),
                "Total Items": format_count(cls.get('total', 0)),
                "Matched Items": format_count(cls.get('matched', 0)),
                "Coverage": format_percentage(cls.get('percent', 0))
            }
            for cls in metrics.classification_coverage
        ]
        exporter.add_kpi_sheet("Classification Coverage", class_data)

    return exporter.save()


def export_review_to_excel(metrics, org_id: str, project_id: str) -> bytes:
    """Export review queue metrics to Excel."""
    exporter = ExcelExporter("Review Queue Report", org_id, project_id)
    exporter.add_metadata_sheet()

    # Summary
    clean_items = metrics.total_pending - metrics.critical_flags_count - metrics.advisory_flags_count
    summary_data = [
        {
            "Metric": "Total Pending",
            "Value": format_count(metrics.total_pending)
        },
        {
            "Metric": "Critical Flags",
            "Value": format_count(metrics.critical_flags_count)
        },
        {
            "Metric": "Advisory Flags",
            "Value": format_count(metrics.advisory_flags_count)
        },
        {
            "Metric": "Clean Items",
            "Value": format_count(clean_items)
        },
        {
            "Metric": "High Urgency",
            "Value": format_count(metrics.high_urgency)
        },
        {
            "Metric": "Medium Urgency",
            "Value": format_count(metrics.medium_urgency)
        },
        {
            "Metric": "Low Urgency",
            "Value": format_count(metrics.low_urgency)
        },
        {
            "Metric": "Oldest Review (days)",
            "Value": f"{metrics.oldest_review_days:.1f}" if metrics.oldest_review_days else "N/A"
        },
        {
            "Metric": "Average Age (days)",
            "Value": f"{metrics.avg_age_days:.1f}" if metrics.avg_age_days else "N/A"
        },
        {
            "Metric": "Items >7 Days",
            "Value": format_count(metrics.items_over_7_days)
        },
        {
            "Metric": "Items >30 Days",
            "Value": format_count(metrics.items_over_30_days)
        }
    ]
    exporter.add_kpi_sheet("Review Summary", summary_data)

    # Confidence Distribution
    confidence_data = [
        {
            "Confidence Level": "High (≥85%)",
            "Count": format_count(metrics.confidence_high)
        },
        {
            "Confidence Level": "Medium (70-84%)",
            "Count": format_count(metrics.confidence_medium)
        },
        {
            "Confidence Level": "Low (<70%)",
            "Count": format_count(metrics.confidence_low)
        }
    ]
    exporter.add_kpi_sheet("Confidence Distribution", confidence_data)

    # Classification Breakdown
    if metrics.classification_breakdown:
        class_data = [
            {
                "Classification Code": cls.get('code', 'N/A'),
                "Total Items": format_count(cls.get('total', 0)),
                "Critical Flags": format_count(cls.get('critical', 0)),
                "Advisory Flags": format_count(cls.get('advisory', 0)),
                "Avg Confidence": format_percentage(cls.get('avg_confidence', 0)) if cls.get('avg_confidence') else "N/A"
            }
            for cls in metrics.classification_breakdown
        ]
        exporter.add_kpi_sheet("Classification Breakdown", class_data)

    return exporter.save()


def export_reports_to_excel(metrics, org_id: str, project_id: str) -> bytes:
    """Export financial reports to Excel."""
    exporter = ExcelExporter("Financial Report", org_id, project_id)
    exporter.add_metadata_sheet()

    # Financial Summary
    vat_amount = metrics.total_cost_gross - metrics.total_cost_net
    cost_at_risk_pct = (metrics.high_risk_total_cost / metrics.total_cost_net * 100) if metrics.total_cost_net > 0 else 0

    summary_data = [
        {
            "Metric": "Total Cost (Net)",
            "Value": format_currency(metrics.total_cost_net, metrics.currency)
        },
        {
            "Metric": "Total Cost (Gross)",
            "Value": format_currency(metrics.total_cost_gross, metrics.currency)
        },
        {
            "Metric": "VAT Amount",
            "Value": format_currency(vat_amount, metrics.currency)
        },
        {
            "Metric": "High Confidence Cost",
            "Value": format_currency(metrics.high_confidence_cost, metrics.currency)
        },
        {
            "Metric": "Medium Confidence Cost",
            "Value": format_currency(metrics.medium_confidence_cost, metrics.currency)
        },
        {
            "Metric": "Low Confidence Cost",
            "Value": format_currency(metrics.low_confidence_cost, metrics.currency)
        },
        {
            "Metric": "High Risk Total Cost",
            "Value": format_currency(metrics.high_risk_total_cost, metrics.currency)
        },
        {
            "Metric": "Cost at Risk %",
            "Value": format_percentage(cost_at_risk_pct)
        },
        {
            "Metric": "Total Items",
            "Value": format_count(metrics.total_items)
        },
        {
            "Metric": "Matched Items",
            "Value": format_count(metrics.matched_items)
        },
        {
            "Metric": "Unmatched Items",
            "Value": format_count(metrics.unmatched_items_count)
        },
        {
            "Metric": "Match Coverage",
            "Value": format_percentage(metrics.match_percentage)
        },
        {
            "Metric": "Average Unit Price",
            "Value": format_currency(metrics.avg_unit_price, metrics.currency) if metrics.avg_unit_price else "N/A"
        },
        {
            "Metric": "Average Confidence",
            "Value": format_percentage(metrics.avg_confidence) if metrics.avg_confidence else "N/A"
        }
    ]
    exporter.add_kpi_sheet("Financial Summary", summary_data)

    # Classification Cost Breakdown
    if hasattr(metrics, 'classification_cost_breakdown') and metrics.classification_cost_breakdown:
        class_cost_data = [
            {
                "Classification": cls.get('code', 'N/A'),
                "Name": cls.get('name', 'N/A'),
                "Items": format_count(cls.get('count', 0)),
                "Net Cost": format_currency(cls.get('net_cost', 0), metrics.currency),
                "Gross Cost": format_currency(cls.get('gross_cost', 0), metrics.currency),
                "Avg Confidence": format_percentage(cls.get('avg_confidence', 0)) if cls.get('avg_confidence') else "N/A"
            }
            for cls in metrics.classification_cost_breakdown
        ]
        exporter.add_kpi_sheet("Cost by Classification", class_cost_data)

    # Top Expensive Items
    if metrics.top_10_expensive:
        top_items_data = [
            {
                "Family": item.get('family', 'N/A')[:50],
                "Type": item.get('type', 'N/A')[:50] if item.get('type') else "N/A",
                "Unit Cost": format_currency(item.get('unit_price', 0), metrics.currency),
                "Total Cost": format_currency(item.get('cost', 0), metrics.currency),
                "Confidence": format_percentage(item.get('confidence', 0))
            }
            for item in metrics.top_10_expensive
        ]
        exporter.add_kpi_sheet("Top 10 Expensive Items", top_items_data)

    return exporter.save()


def export_audit_to_excel(metrics, org_id: str, project_id: str) -> bytes:
    """Export audit trail metrics to Excel."""
    exporter = ExcelExporter("Audit Trail Report", org_id, project_id)
    exporter.add_metadata_sheet()

    # Summary
    high_confidence_pct = (metrics.high_confidence_count / metrics.total_decisions * 100) if metrics.total_decisions > 0 else 0

    summary_data = [
        {
            "Metric": "Total Decisions",
            "Value": format_count(metrics.total_decisions)
        },
        {
            "Metric": "Total Items Audited",
            "Value": format_count(metrics.total_items_audited)
        },
        {
            "Metric": "Auto-Approved",
            "Value": format_count(metrics.auto_approved_count)
        },
        {
            "Metric": "User Approved",
            "Value": format_count(metrics.user_approved_count)
        },
        {
            "Metric": "Manual Review",
            "Value": format_count(metrics.manual_review_count)
        },
        {
            "Metric": "Rejected",
            "Value": format_count(metrics.rejected_count)
        },
        {
            "Metric": "System Automation %",
            "Value": format_percentage(metrics.system_percentage)
        },
        {
            "Metric": "High Confidence %",
            "Value": format_percentage(high_confidence_pct)
        },
        {
            "Metric": "Average Confidence",
            "Value": format_percentage(metrics.avg_confidence) if metrics.avg_confidence else "N/A"
        },
        {
            "Metric": "Compliance Score",
            "Value": f"{metrics.compliance_score}/100 ({metrics.compliance_status})"
        },
        {
            "Metric": "Decisions (Last 7 Days)",
            "Value": format_count(metrics.decisions_last_7_days)
        },
        {
            "Metric": "Decisions (Last 30 Days)",
            "Value": format_count(metrics.decisions_last_30_days)
        },
        {
            "Metric": "Avg Decisions/Day",
            "Value": f"{metrics.avg_decisions_per_day:.1f}"
        },
        {
            "Metric": "Peak Decision Day",
            "Value": metrics.peak_decision_day if metrics.peak_decision_day else "N/A"
        },
        {
            "Metric": "Peak Decision Count",
            "Value": format_count(metrics.peak_decision_count)
        }
    ]
    exporter.add_kpi_sheet("Audit Summary", summary_data)

    # Confidence Distribution
    confidence_data = [
        {
            "Confidence Level": "High (≥85%)",
            "Count": format_count(metrics.high_confidence_count)
        },
        {
            "Confidence Level": "Medium (70-84%)",
            "Value": format_count(metrics.medium_confidence_count)
        },
        {
            "Confidence Level": "Low (<70%)",
            "Count": format_count(metrics.low_confidence_count)
        }
    ]
    exporter.add_kpi_sheet("Confidence Distribution", confidence_data)

    # Decision Sources
    sources_data = [
        {
            "Source": "Mapping Memory",
            "Count": format_count(metrics.mapping_memory_count)
        },
        {
            "Source": "Fuzzy Match",
            "Count": format_count(metrics.fuzzy_match_count)
        },
        {
            "Source": "Review UI",
            "Count": format_count(metrics.review_ui_count)
        }
    ]
    exporter.add_kpi_sheet("Decision Sources", sources_data)

    # Daily Timeline
    if metrics.daily_timeline:
        timeline_data = [
            {
                "Date": entry.get('date', 'N/A'),
                "Decisions": format_count(entry.get('count', 0)),
                "Avg Confidence": format_percentage(entry.get('avg_confidence', 0)) if entry.get('avg_confidence') else "N/A"
            }
            for entry in metrics.daily_timeline
        ]
        exporter.add_kpi_sheet("Daily Timeline", timeline_data)

    # Top Reviewers
    if metrics.top_reviewers:
        reviewers_data = [
            {
                "Reviewer": reviewer.get('created_by', 'N/A'),
                "Decisions": format_count(reviewer.get('count', 0)),
                "Avg Confidence": format_percentage(reviewer.get('avg_confidence', 0)) if reviewer.get('avg_confidence') else "N/A"
            }
            for reviewer in metrics.top_reviewers
        ]
        exporter.add_kpi_sheet("Top Reviewers", reviewers_data)

    return exporter.save()


def export_prices_to_excel(metrics, org_id: str) -> bytes:
    """Export price data quality metrics to Excel."""
    exporter = ExcelExporter("Price Data Quality Report", org_id, "All Projects")
    exporter.add_metadata_sheet()

    # Summary
    current_pct = (metrics.current_price_items / metrics.total_price_items * 100) if metrics.total_price_items > 0 else 0
    summary_data = [
        {
            "Metric": "Total Price Items",
            "Value": format_count(metrics.total_price_items)
        },
        {
            "Metric": "Current Price Items",
            "Value": format_count(metrics.current_price_items)
        },
        {
            "Metric": "Current Items %",
            "Value": format_percentage(current_pct)
        },
        {
            "Metric": "Historical Price Items",
            "Value": format_count(metrics.historical_price_items)
        },
        {
            "Metric": "Primary Currency",
            "Value": metrics.currency
        },
        {
            "Metric": "Unique Vendors",
            "Value": format_count(metrics.unique_vendors)
        },
        {
            "Metric": "Min Unit Price",
            "Value": format_currency(metrics.min_unit_price, metrics.currency) if metrics.min_unit_price else "N/A"
        },
        {
            "Metric": "Max Unit Price",
            "Value": format_currency(metrics.max_unit_price, metrics.currency) if metrics.max_unit_price else "N/A"
        },
        {
            "Metric": "Avg Unit Price",
            "Value": format_currency(metrics.avg_unit_price, metrics.currency) if metrics.avg_unit_price else "N/A"
        },
        {
            "Metric": "Median Unit Price",
            "Value": format_currency(metrics.median_unit_price, metrics.currency) if metrics.median_unit_price else "N/A"
        },
        {
            "Metric": "Classification Coverage",
            "Value": format_percentage(metrics.classification_coverage_pct)
        },
        {
            "Metric": "Classifications with Prices",
            "Value": format_count(metrics.classifications_with_prices)
        },
        {
            "Metric": "Quality Score",
            "Value": f"{metrics.quality_score}/100 ({metrics.quality_status})"
        },
        {
            "Metric": "Avg Age (days)",
            "Value": f"{metrics.avg_age_days:.1f}" if metrics.avg_age_days else "N/A"
        },
        {
            "Metric": "Oldest Price (days)",
            "Value": format_count(metrics.oldest_price_days) if metrics.oldest_price_days else "N/A"
        },
        {
            "Metric": "Updated Last 30 Days",
            "Value": format_count(metrics.prices_updated_last_30_days)
        },
        {
            "Metric": "Updated Last 90 Days",
            "Value": format_count(metrics.prices_updated_last_90_days)
        },
        {
            "Metric": "Stale Prices (>1 year)",
            "Value": format_count(metrics.stale_prices_count)
        },
        {
            "Metric": "VAT Specified Count",
            "Value": format_count(metrics.vat_specified_count)
        },
        {
            "Metric": "VAT Unspecified Count",
            "Value": format_count(metrics.vat_unspecified_count)
        }
    ]
    exporter.add_kpi_sheet("Price Summary", summary_data)

    # Top Classifications
    if metrics.top_classifications:
        class_data = [
            {
                "Classification Code": cls.get('code', 'N/A'),
                "Count": format_count(cls.get('count', 0)),
                "Avg Price": format_currency(cls.get('avg_price', 0), metrics.currency),
                "Min Price": format_currency(cls.get('min_price', 0), metrics.currency),
                "Max Price": format_currency(cls.get('max_price', 0), metrics.currency)
            }
            for cls in metrics.top_classifications
        ]
        exporter.add_kpi_sheet("Top Classifications", class_data)

    # Top Expensive Items
    if metrics.top_10_expensive:
        expensive_data = [
            {
                "Item Code": item.get('code', 'N/A'),
                "Description": item.get('description', 'N/A')[:100],
                "Unit Price": format_currency(item.get('unit_price', 0), metrics.currency),
                "Vendor": item.get('vendor', 'Unknown'),
                "Last Updated": item.get('updated', 'N/A')
            }
            for item in metrics.top_10_expensive
        ]
        exporter.add_kpi_sheet("Most Expensive Items", expensive_data)

    # Top Vendors
    if metrics.top_vendors:
        vendor_data = [
            {
                "Vendor": vendor.get('vendor', 'N/A'),
                "Price Items": format_count(vendor.get('count', 0)),
                "Avg Unit Price": format_currency(vendor.get('avg_price', 0), metrics.currency)
            }
            for vendor in metrics.top_vendors
        ]
        exporter.add_kpi_sheet("Top Vendors", vendor_data)

    return exporter.save()


def export_to_csv(data: list[dict[str, Any]], title: str) -> str:
    """Export data to CSV format."""
    if not data:
        return f"{title}\nNo data available"

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)

    return output.getvalue()
