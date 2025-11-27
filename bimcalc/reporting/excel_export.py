"""Excel export functionality for BIMCalc projects.

Generates comprehensive Excel workbooks with:
- Cost breakdown by category
- Category-specific labor analysis
- Matched items list
- Charts and visualizations
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from sqlalchemy.ext.asyncio import AsyncSession
    from bimcalc.reporting.dashboard_metrics import DashboardMetrics


async def generate_cost_breakdown_excel(
    session: AsyncSession,
    org_id: str,
    project_id: str
) -> BytesIO:
    """Generate Excel workbook with comprehensive cost breakdown.
    
    Args:
        session: Database session
        org_id: Organization ID
        project_id: Project ID
        
    Returns:
        BytesIO containing Excel workbook
    """
    from bimcalc.reporting.dashboard_metrics import compute_dashboard_metrics
    from bimcalc.db.models import ProjectModel, LaborRateOverride
    from sqlalchemy import select
    
    # Get project info
    project_query = select(ProjectModel).where(
        ProjectModel.org_id == org_id,
        ProjectModel.project_id == project_id
    )
    project = (await session.execute(project_query)).scalar_one_or_none()
    
    if not project:
        raise ValueError(f"Project not found: {org_id}/{project_id}")
    
    # Compute dashboard metrics
    metrics = await compute_dashboard_metrics(session, org_id, project_id)
    
    # Get labor rate overrides
    labor_rates_query = select(LaborRateOverride).where(
        LaborRateOverride.project_id == project.id
    )
    labor_overrides = (await session.execute(labor_rates_query)).scalars().all()
    
    base_rate = project.settings.get('blended_labor_rate', 50.0) if project.settings else 50.0
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets
    _create_summary_sheet(wb, project, metrics)
    _create_category_analysis_sheet(wb, metrics, labor_overrides, base_rate)
    await _create_items_sheet(wb, session, org_id, project_id)
    
    # Add charts to summary
    await _add_cost_charts(wb, metrics)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def _create_summary_sheet(wb: Workbook, project, metrics: DashboardMetrics):
    """Create cost summary sheet with key metrics."""
    ws = wb.create_sheet("Cost Summary", 0)
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Title
    ws['A1'] = "BIMCalc Cost Breakdown Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    # Project info
    ws['A3'] = "Project:"
    ws['B3'] = project.display_name
    ws['A4'] = "Organization:"
    ws['B4'] = project.org_id
    ws['A5'] = "Generated:"
    from datetime import datetime
    ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    # Metrics section
    ws['A7'] = "Financial Summary"
    ws['A7'].font = Font(bold=True, size=14)
    
    # Headers
    row = 9
    headers = ["Metric", "Value", "Currency"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    data = [
        ("Total Material Cost", f"€{metrics.total_cost_net:,.2f}", metrics.currency),
        ("Total Material Cost (with markup)", f"€{metrics.total_cost_net * 1.15:,.2f}", metrics.currency),
        ("Total Labor Hours", f"{metrics.total_labor_hours:,.1f} hrs", "-"),
        ("Total Labor Cost", f"€{metrics.total_labor_cost:,.2f}", metrics.currency),
        ("Total Installed Cost", f"€{metrics.total_installed_cost:,.2f}", metrics.currency),
        ("", "", ""),
        ("Total Items", str(metrics.total_items), "-"),
        ("Matched Items", str(metrics.matched_items), "-"),
        ("Match Rate", f"{metrics.match_percentage:.1f}%", "-"),
    ]
    
    for row_data in data:
        row += 1
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12


def _create_category_analysis_sheet(
    wb: Workbook,
    metrics: DashboardMetrics,
    labor_overrides: list,
    base_rate: float
):
    """Create category-specific labor analysis sheet."""
    ws = wb.create_sheet("Category Labor Rates")
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Title
    ws['A1'] = "Category-Specific Labor Rates"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Base rate
    ws['A3'] = "Base Labor Rate:"
    ws['B3'] = f"€{base_rate:.2f}/hr"
    ws['B3'].font = Font(bold=True)
    
    # Category overrides table
    ws['A5'] = "Category Overrides"
    ws['A5'].font = Font(bold=True, size=12)
    
    # Headers
    row = 7
    headers = ["Category", "Rate (€/hr)", "vs Base"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    if labor_overrides:
        for override in labor_overrides:
            row += 1
            ws.cell(row=row, column=1, value=override.category)
            ws.cell(row=row, column=2, value=f"€{float(override.rate):.2f}")
            diff = float(override.rate) - base_rate
            ws.cell(row=row, column=3, value=f"{'+' if diff > 0 else ''}{diff:.2f}")
    else:
        row += 1
        ws.cell(row=row, column=1, value="No category overrides defined")
        ws.merge_cells(f'A{row}:C{row}')
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12


async def _create_items_sheet(
    wb: Workbook,
    session: AsyncSession,
    org_id: str,
    project_id: str
):
    """Create items list with pricing."""
    from sqlalchemy import select, text
    
    ws = wb.create_sheet("Items List")
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Title
    ws['A1'] = "Matched Items"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    row = 3
    headers = ["Category", "Description", "Quantity", "Unit", "Unit Price", "Total Cost", "Labor Hours"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Query items
    query = text("""
        WITH latest_matches AS (
            SELECT 
                mr.item_id,
                mr.price_item_id,
                ROW_NUMBER() OVER (PARTITION BY mr.item_id ORDER BY mr.timestamp DESC) as rn
            FROM match_results mr
            JOIN items i ON i.id = mr.item_id
            WHERE i.org_id = :org_id
              AND i.project_id = :project_id
              AND mr.decision IN ('auto-accepted', 'accepted', 'pending-review')
        ),
        active_mappings AS (
            SELECT canonical_key, price_item_id
            FROM item_mapping
            WHERE org_id = :org_id AND end_ts IS NULL
        )
        SELECT
            i.category,
            i.type_name as description,
            i.quantity,
            pi.unit,
            pi.unit_price,
            pi.labor_hours
        FROM items i
        LEFT JOIN latest_matches lm ON lm.item_id = i.id AND lm.rn = 1
        LEFT JOIN active_mappings am ON am.canonical_key = i.canonical_key
        LEFT JOIN price_items pi ON pi.id = COALESCE(lm.price_item_id, am.price_item_id)
        WHERE i.org_id = :org_id
          AND i.project_id = :project_id
        ORDER BY i.category, i.type_name
        LIMIT 1000
    """)
    
    result = await session.execute(query, {"org_id": org_id, "project_id": project_id})
    
    # Data rows
    for item in result:
        row += 1
        ws.cell(row=row, column=1, value=item.category or "Uncategorized")
        ws.cell(row=row, column=2, value=item.description)
        ws.cell(row=row, column=3, value=float(item.quantity) if item.quantity else 0)
        ws.cell(row=row, column=4, value=item.unit or "-")
        
        if item.unit_price:
            ws.cell(row=row, column=5, value=float(item.unit_price))
            total_cost = float(item.quantity * item.unit_price) if item.quantity and item.unit_price else 0
            ws.cell(row=row, column=6, value=total_cost)
        else:
            ws.cell(row=row, column=5, value="-")
            ws.cell(row=row, column=6, value="-")
        
        labor_hours = float(item.labor_hours) if item.labor_hours else 0
        ws.cell(row=row, column=7, value=labor_hours if labor_hours > 0 else "-")
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12


async def _add_cost_charts(wb: Workbook, metrics: DashboardMetrics):
    """Add charts to summary sheet."""
    ws = wb["Cost Summary"]
    
    # Simple placeholder - actual chart implementation
    # Charts are complex in openpyxl, so starting with a note
    ws['F7'] = "Cost Distribution Chart"
    ws['F7'].font = Font(bold=True, size=12)
    ws['F9'] = "Material Cost:"
    ws['G9'] = f"€{metrics.total_cost_net:,.2f}"
    ws['F10'] = "Labor Cost:"
    ws['G10'] = f"€{metrics.total_labor_cost:,.2f}"
    
    # TODO: Add actual pie chart for cost distribution
    # PieChart implementation can be added in next iteration
