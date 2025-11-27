"""Export utilities for Intelligence features."""

import csv
from io import BytesIO, StringIO
from datetime import datetime
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus.flowables import HRFlowable
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, LineChart, BarChart, Reference

from bimcalc.db.models import QAChecklistModel, ItemModel


class ChecklistPDFExporter:
    """Export QA checklists to professional PDF format."""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Add custom paragraph styles."""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor("#1a365d"),
            spaceAfter=30,
            alignment=1  # Center
        ))
        
        # Subtitle
        self.styles.add(ParagraphStyle(
            name='Subtitle',
            parent=self.styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor("#4a5568"),
            spaceAfter=20,
            alignment=1
        ))
        
        # Category header
        self.styles.add(ParagraphStyle(
            name='CategoryHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor("#2d3748"),
            spaceBefore=15,
            spaceAfter=10
        ))
    
    def export(self, checklist: QAChecklistModel, item: ItemModel) -> bytes:
        """Generate PDF for checklist.
        
        Args:
            checklist: QA checklist to export
            item: Associated item
            
        Returns:
            PDF file as bytes
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                               rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=18)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Title
        elements.append(Paragraph("QA Testing Checklist", self.styles['CustomTitle']))
        
        # Item details
        item_info = f"""
        <b>Item:</b> {item.family} - {item.type_name}<br/>
        <b>Classification:</b> {item.classification_code or 'N/A'}<br/>
        <b>Generated:</b> {checklist.generated_at.strftime('%Y-%m-%d %H:%M')}<br/>
        <b>Completion:</b> {checklist.completion_percent:.0f}%
        """
        elements.append(Paragraph(item_info, self.styles['Subtitle']))
        elements.append(Spacer(1, 20))
        
        # Progress bar (visual)
        elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#e2e8f0")))
        elements.append(Spacer(1, 10))
        
        # Checklist items grouped by category
        items = checklist.checklist_items.get("items", [])
        categories = {}
        for item_data in items:
            cat = item_data.get("category", "General")
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(item_data)
        
        for category, cat_items in categories.items():
            # Category header
            elements.append(Paragraph(f"✓ {category}", self.styles['CategoryHeader']))
            
            # Table for items
            table_data = [["", "Requirement", "Priority", "Time"]]
            
            for item_data in cat_items:
                checkbox = "☑" if item_data.get("completed") else "☐"
                req = item_data.get("requirement", "")
                priority = item_data.get("priority", "Medium")
                time = f"{item_data.get('estimated_time_minutes', 10)} min"
                
                table_data.append([checkbox, req, priority, time])
            
            # Create table
            table = Table(table_data, colWidths=[0.5*inch, 4*inch, 1*inch, 0.8*inch])
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f7fafc")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#2d3748")),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Body
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor("#4a5568")),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 15))
        
        # Source documents
        if checklist.source_documents.get("docs"):
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("Source Documents", self.styles['CategoryHeader']))
            
            docs_text = "<br/>".join([
                f"• {doc['title']}" 
                for doc in checklist.source_documents.get("docs", [])
            ])
            elements.append(Paragraph(docs_text, self.styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer and return it
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes


class AnalyticsExcelExporter:
    """Export analytics data to Excel workbook."""
    
    def export(self, analytics_data: dict[str, Any]) -> bytes:
        """Generate Excel workbook with analytics.
        
        Args:
            analytics_data: Dict with keys:
                - classification_breakdown
                - compliance_timeline
                - cost_distribution
                - document_coverage
        
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Classification Breakdown
        self._create_classification_sheet(wb, analytics_data.get("classification_breakdown", {}))
        
        # 2. Compliance Timeline
        self._create_compliance_sheet(wb, analytics_data.get("compliance_timeline", {}))
        
        # 3. Cost Distribution
        self._create_cost_sheet(wb, analytics_data.get("cost_distribution", {}))
        
        # 4. Document Coverage
        self._create_coverage_sheet(wb, analytics_data.get("document_coverage", {}))
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_classification_sheet(self, wb: Workbook, data: dict):
        """Create classification breakdown sheet."""
        ws = wb.create_sheet("Classification Breakdown")
        
        # Headers
        ws['A1'] = "Classification"
        ws['B1'] = "Item Count"
        self._style_header(ws, 'A1:B1')
        
        # Data
        labels = data.get("labels", [])
        values = data.get("values", [])
        
        for idx, (label, value) in enumerate(zip(labels, values), start=2):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
        
        # Chart
        chart = PieChart()
        chart.title = "Items by Classification"
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(values)+1)
        labels_ref = Reference(ws, min_col=1, min_row=2, max_row=len(values)+1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        ws.add_chart(chart, "D2")
    
    def _create_compliance_sheet(self, wb: Workbook, data: dict):
        """Create compliance timeline sheet."""
        ws = wb.create_sheet("Compliance Trend")
        
        # Headers
        ws['A1'] = "Date"
        ws['B1'] = "Compliance %"
        self._style_header(ws, 'A1:B1')
        
        # Data
        dates = data.get("dates", [])
        percentages = data.get("percentages", [])
        
        for idx, (date, pct) in enumerate(zip(dates, percentages), start=2):
            ws[f'A{idx}'] = date
            ws[f'B{idx}'] = pct
        
        # Chart
        chart = LineChart()
        chart.title = "QA Compliance Over Time"
        chart.y_axis.title = "Completion %"
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(dates)+1)
        labels_ref = Reference(ws, min_col=1, min_row=2, max_row=len(dates)+1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        ws.add_chart(chart, "D2")
    
    def _create_cost_sheet(self, wb: Workbook, data: dict):
        """Create cost distribution sheet."""
        ws = wb.create_sheet("Cost Distribution")
        
        # Headers
        ws['A1'] = "Classification"
        ws['B1'] = "Total Cost"
        self._style_header(ws, 'A1:B1')
        
        # Data
        labels = data.get("labels", [])
        values = data.get("values", [])
        
        for idx, (label, value) in enumerate(zip(labels, values), start=2):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
            ws[f'B{idx}'].number_format = '$#,##0.00'
        
        # Chart
        chart = BarChart()
        chart.title = "Cost by Classification"
        chart.y_axis.title = "Cost ($)"
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(values)+1)
        labels_ref = Reference(ws, min_col=1, min_row=2, max_row=len(values)+1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        ws.add_chart(chart, "D2")
    
    def _create_coverage_sheet(self, wb: Workbook, data: dict):
        """Create document coverage matrix sheet."""
        ws = wb.create_sheet("Document Coverage")
        
        matrix = data.get("matrix", [])
        row_labels = data.get("row_labels", [])
        col_labels = data.get("col_labels", [])
        
        # Column headers (classifications)
        for col_idx, label in enumerate(col_labels, start=2):
            ws.cell(1, col_idx, label)
        
        # Row headers (document types) + data
        for row_idx, (row_label, row_data) in enumerate(zip(row_labels, matrix), start=2):
            ws.cell(row_idx, 1, row_label)
            for col_idx, value in enumerate(row_data, start=2):
                cell = ws.cell(row_idx, col_idx, value)
                # Color code based on value
                if value > 5:
                    cell.fill = PatternFill(start_color="d1fae5", fill_type="solid")
                elif value > 0:
                    cell.fill = PatternFill(start_color="fef3c7", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="fee2e2", fill_type="solid")
        
        self._style_header(ws, f'A1:{chr(65+len(col_labels))}1')
    
    def _style_header(self, ws, range_str: str):
        """Apply header styling to range."""
        for row in ws[range_str]:
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2d3748", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")


class RiskCSVExporter:
    """Export risk assessment data to CSV."""
    
    def export(self, risk_items: list[dict]) -> str:
        """Generate CSV of risk data.
        
        Args:
            risk_items: List of items with risk scores
            
        Returns:
            CSV string
        """
        output = StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            "Item ID",
            "Family",
            "Type",
            "Classification",
            "Risk Score",
            "Risk Level",
            "Top Recommendation 1",
            "Top Recommendation 2"
        ])
        
        # Data rows
        for item in risk_items:
            recommendations = item.get("recommendations", [])
            rec1 = recommendations[0] if len(recommendations) > 0 else ""
            rec2 = recommendations[1] if len(recommendations) > 1 else ""
            
            writer.writerow([
                item.get("item_id", ""),
                item.get("family", ""),
                item.get("type_name", ""),
                item.get("classification_code", ""),
                item.get("risk_score", 0),
                item.get("risk_level", ""),
                rec1,
                rec2
            ])
        
        csv_string = output.getvalue()
        output.close()
        return csv_string
