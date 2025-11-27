"""Verification script for Excel export functionality."""
import asyncio
import os
from io import BytesIO
from uuid import uuid4
from decimal import Decimal

from openpyxl import load_workbook
from bimcalc.reporting.excel_export import generate_cost_breakdown_excel
from bimcalc.db.connection import get_session
from bimcalc.db.models import ProjectModel, ItemModel, LaborRateOverride

async def verify_excel_export():
    print("🧪 Verifying Excel Export...")
    
    async with get_session() as session:
        # 1. Create test data
        org_id = "test-org-export"
        project_id = f"proj-{uuid4().hex[:8]}"
        
        print(f"   Creating test project: {project_id}")
        
        project = ProjectModel(
            id=uuid4(),
            org_id=org_id,
            project_id=project_id,
            display_name="Export Test Project",
            status="active",
            settings={"blended_labor_rate": 60.0}
        )
        session.add(project)
        
        # Add labor override
        override = LaborRateOverride(
            id=uuid4(),
            project_id=project.id,
            category="Electrical",
            rate=Decimal("75.00")
        )
        session.add(override)
        
        # Add some items (simplified for test)
        item = ItemModel(
            id=uuid4(),
            org_id=org_id,
            project_id=project_id,
            category="Electrical",
            family="Test Family",
            type_name="Test Type",
            quantity=Decimal("10.0")
        )
        session.add(item)
        
        await session.commit()
        
        try:
            # 2. Generate Excel
            print("   Generating Excel workbook...")
            excel_bytes = await generate_cost_breakdown_excel(session, org_id, project_id)
            
            # 3. Verify content
            print("   Verifying workbook content...")
            wb = load_workbook(excel_bytes)
            
            # Check sheets
            expected_sheets = ["Cost Summary", "Category Labor Rates", "Items List"]
            for sheet in expected_sheets:
                if sheet in wb.sheetnames:
                    print(f"   ✅ Sheet found: {sheet}")
                else:
                    print(f"   ❌ Missing sheet: {sheet}")
                    return
            
            # Check Category Labor Rates content
            ws_cats = wb["Category Labor Rates"]
            found_override = False
            for row in ws_cats.iter_rows(values_only=True):
                if row[0] == "Electrical" and row[1] == "€75.00":
                    found_override = True
                    break
            
            if found_override:
                print("   ✅ Category override data verified")
            else:
                print("   ❌ Category override data missing or incorrect")
            
            # Save for manual inspection
            filename = "verify_export.xlsx"
            with open(filename, "wb") as f:
                f.write(excel_bytes.getvalue())
            print(f"   💾 Saved test file to: {filename}")
            
        finally:
            # Cleanup
            print("   Cleaning up test data...")
            await session.delete(item)
            await session.delete(override)
            await session.delete(project)
            await session.commit()

if __name__ == "__main__":
    asyncio.run(verify_excel_export())
